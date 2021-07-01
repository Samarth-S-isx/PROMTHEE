import numpy as np
import openpyxl
from openpyxl import Workbook

# Loading Worksheets
wb = openpyxl.load_workbook(
    "C:\\Users\\SAMARTH_IITKGP\\Tre\\19CE10058_SAMARTHSAXENA.xlsx")
sh2 = wb['Sheet2']
wbr = Workbook()
shr = wbr.active
wbr['Sheet'].title = "Result"


# Creating Decision matrix using data from excel
elements_list = []
row = sh2.max_row
column = sh2.max_column
for r in range(3, row):
    for c in range(2, column+1):
        elements_list.append(float((sh2.cell(r, c).value)))
matrix = np.array(elements_list).reshape(row-3, column-1)


# weight matrix creation using data from excel
weight_list = []
for c in range(2, column+1):
    weight_list.append(float((sh2.cell(row, c).value)))
Weight_matrix = np.array(weight_list).reshape(1, column-1)


# min-max for each category
min_matrix = np.min(matrix, axis=0)
max_matrix = np.max(matrix, axis=0)
diff_matrix = max_matrix-min_matrix

# Knowing the size of desicion matrix
R = len(matrix)
C = len(matrix[0])


# # Step3 Normalizing the descision matrix
for c in range(2, column+1):
    criteria = str(sh2.cell(2, c).value)
    if criteria.lower() == "cost":
        for r in range(0, R):
            matrix[r][c-2] = ((max_matrix[c-2]-matrix[r][c-2])
                              )/diff_matrix[c-2]
    else:
        for r in range(0, R):
            matrix[r][c-2] = (matrix[r][c-2]-min_matrix[c-2])/diff_matrix[c-2]

# print(matrix)

# Step 4 determination of deviation by pairwise comparision
step2_diff = []
for c in range(0, C):
    for r in range(R):
        for i in range(R):
            if r == i:
                continue
            else:
                diff2 = matrix[r][c]-matrix[i][c]
                step2_diff.append(diff2)
# Step 5 Aplying Prefrence Function making neg elemnt 0
for i in range(len(step2_diff)):
    if step2_diff[i] < 0:
        step2_diff[i] = 0
    else:
        continue
Diff2_matrix = np.array(step2_diff).reshape(C, int(len(step2_diff)/C))
# Converting Pairwise deviation into matrix form
Pairwise_deviation_matrix = Diff2_matrix.transpose()

# Step 6 to calculate agregate function using weight
Sum_of_weight = np.sum(Weight_matrix, axis=1)
Weight_matrix = Pairwise_deviation_matrix*Weight_matrix
Sum_of_row = np.sum(Weight_matrix, axis=1)
New_list = []
for sum in Sum_of_row:
    New_list.append(sum)

for i in range(C):
    New_list.insert((C+1)*i, 0)
# Aggregate preference function
Aggregte_Matrix = np.array(New_list).reshape(C, C)

# Step 7 Calculating entering flow and outgoing flow
Leaving_Flow = np.sum(Aggregte_Matrix, axis=1)/(C-1)
Entering_Flow = np.sum(Aggregte_Matrix, axis=0)/(C-1)
# Step 8 Calculating net flow
Leaving_Entering = Leaving_Flow-Entering_Flow

# Step 9 Ranking them according to there rank
Preference = []
for i in range(len(Leaving_Entering)):
    Preference.append((Leaving_Entering[i], sh2.cell(i+3, 1).value))
Preference = sorted(Preference, reverse=True)
print(Preference)

# showing result in new worksheet
for i in range(C):
    for j in range(2):
        shr.cell(row=1+i, column=1+j).value = Preference[i][j]
wbr.save("C:\\Users\\SAMARTH_IITKGP\\Tre\\Result.xlsx")
